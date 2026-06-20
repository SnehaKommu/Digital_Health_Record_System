from flask import Flask, render_template, request, redirect, session, flash, send_from_directory, send_file
from db_config import get_db
import os
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR  = os.path.dirname(BASE_DIR)
FRONTEND_DIR = os.path.join(PROJECT_DIR, "frontend")
UPLOAD_DIR   = os.path.join(PROJECT_DIR, "uploads")

app = Flask(
    __name__,
    template_folder=os.path.join(FRONTEND_DIR, "templates"),
    static_folder=os.path.join(FRONTEND_DIR, "static"),
)
app.secret_key = "health_secret_2025"
app.config["UPLOAD_FOLDER"] = UPLOAD_DIR
os.makedirs(UPLOAD_DIR, exist_ok=True)

# ── Auth guard ─────────────────────────────────────────────────────────────────
def require_login(role=None):
    if "user" not in session:
        flash("Please log in to continue.", "danger")
        return redirect("/")
    if role and session.get("role") != role:
        flash("Access denied.", "danger")
        return redirect("/")
    return None

# ══════════════════════════════════════════════════════════════════════════════
# HOME
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/")
def home():
    return render_template("index.html")

# ══════════════════════════════════════════════════════════════════════════════
# LOGIN
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/login", methods=["GET", "POST"])
def login():
    role = request.args.get("role", "worker")

    if request.method == "POST":
        username = request.form["username"].strip()
        password = request.form["password"].strip()
        role     = request.form["role"]

        con = get_db()
        cur = con.cursor()

        if role == "worker":
            cur.execute(
                "SELECT id, name FROM workers WHERE phone=%s AND worker_password=%s",
                (username, password)
            )
            worker = cur.fetchone()
            if worker:
                session["user"]      = worker[1]
                session["role"]      = "worker"
                session["worker_id"] = worker[0]
                flash(f"Welcome, {worker[1]}!", "success")
                return redirect("/worker_dashboard")
            else:
                flash("Invalid phone number or password.", "danger")

        elif role == "doctor":
            cur.execute(
                "SELECT id, name, status FROM doctors WHERE username=%s AND password=%s",
                (username, password)
            )
            doc = cur.fetchone()
            if doc:
                if doc[2] == "pending":
                    flash("Your account is pending admin approval. Please wait.", "warning")
                else:
                    session["user"]      = doc[1]
                    session["role"]      = "doctor"
                    session["doctor_id"] = doc[0]
                    flash(f"Welcome, {doc[1]}!", "success")
                    return redirect("/doctor_dashboard")
            else:
                flash("Invalid username or password.", "danger")

        else:  # admin
            cur.execute(
                "SELECT id, username FROM users WHERE username=%s AND password=%s AND role=%s",
                (username, password, role)
            )
            user = cur.fetchone()
            if user:
                session["user"] = user[1]
                session["role"] = role
                flash("Login successful! Welcome back.", "success")
                return redirect("/admin_dashboard")
            else:
                flash("Invalid credentials. Please try again.", "danger")

    return render_template("login.html", role=role)

# ══════════════════════════════════════════════════════════════════════════════
# DOCTOR SELF-REGISTRATION
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/doctor_register", methods=["GET", "POST"])
def doctor_register():
    if request.method == "POST":
        name           = request.form["name"].strip()
        specialization = request.form["specialization"].strip()
        qualification  = request.form["qualification"].strip()
        hospital       = request.form["hospital"].strip()
        district       = request.form["district"]
        phone          = request.form["phone"].strip()
        email          = request.form["email"].strip()
        experience     = request.form["experience"]
        username       = request.form["username"].strip()
        password       = request.form["password"].strip()
        confirm        = request.form["confirm_password"].strip()

        if password != confirm:
            flash("Passwords do not match.", "danger")
            return redirect("/doctor_register")

        if len(password) < 6:
            flash("Password must be at least 6 characters.", "danger")
            return redirect("/doctor_register")

        con = get_db()
        cur = con.cursor()

        # Check duplicate username
        cur.execute("SELECT id FROM doctors WHERE username=%s", (username,))
        if cur.fetchone():
            flash("Username already taken. Please choose another.", "danger")
            return redirect("/doctor_register")

        # Check duplicate phone
        cur.execute("SELECT id FROM doctors WHERE phone=%s", (phone,))
        if cur.fetchone():
            flash("A doctor with this phone number already exists.", "danger")
            return redirect("/doctor_register")

        cur.execute(
            """INSERT INTO doctors
               (name, specialization, qualification, hospital, district,
                phone, email, experience, username, password, status, source)
               VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'pending','self')""",
            (name, specialization, qualification, hospital, district,
             phone, email, experience, username, password)
        )
        con.commit()
        flash("Registration submitted! Please wait for admin approval before logging in.", "success")
        return redirect("/login?role=doctor")

    return render_template("doctor_register.html")

# ══════════════════════════════════════════════════════════════════════════════
# LOGOUT
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/logout")
def logout():
    username = session.get("user", "")
    role     = session.get("role", "")
    session.clear()
    return render_template("logged_out.html", username=username, role=role)

# ══════════════════════════════════════════════════════════════════════════════
# ADMIN DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/admin_dashboard")
def admin_dashboard():
    guard = require_login("admin")
    if guard: return guard

    con = get_db()
    cur = con.cursor()

    cur.execute("SELECT COUNT(*) FROM workers")
    total_workers = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM records")
    total_records = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM reports")
    total_reports = cur.fetchone()[0]

    cur.execute("SELECT COUNT(DISTINCT region) FROM workers")
    regions = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM doctors WHERE status='approved'")
    total_doctors = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM doctors WHERE status='pending'")
    pending_doctors = cur.fetchone()[0]

    cur.execute("SELECT * FROM workers ORDER BY id DESC LIMIT 5")
    recent_workers = cur.fetchall()

    return render_template(
        "admin_dashboard.html",
        workers=total_workers,
        records=total_records,
        reports=total_reports,
        regions=regions,
        total_doctors=total_doctors,
        pending_doctors=pending_doctors,
        recent_workers=recent_workers
    )

# ══════════════════════════════════════════════════════════════════════════════
# ALL DOCTORS  (admin view)
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/all_doctors")
def all_doctors():
    guard = require_login("admin")
    if guard: return guard

    con = get_db()
    cur = con.cursor()
    cur.execute("SELECT * FROM doctors WHERE status='approved' ORDER BY id")
    doctors = cur.fetchall()

    return render_template("all_doctors.html", doctors=doctors)

# ══════════════════════════════════════════════════════════════════════════════
# PENDING DOCTORS  (admin approves / rejects self-registered doctors)
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/pending_doctors")
def pending_doctors():
    guard = require_login("admin")
    if guard: return guard

    con = get_db()
    cur = con.cursor()
    cur.execute("SELECT * FROM doctors WHERE status='pending' ORDER BY id")
    doctors = cur.fetchall()

    return render_template("pending_doctors.html", doctors=doctors)

@app.route("/approve_doctor/<int:id>")
def approve_doctor(id):
    guard = require_login("admin")
    if guard: return guard

    con = get_db()
    cur = con.cursor()
    cur.execute("UPDATE doctors SET status='approved' WHERE id=%s", (id,))
    con.commit()
    flash("Doctor approved successfully! They can now log in.", "success")
    return redirect("/pending_doctors")

@app.route("/reject_doctor/<int:id>")
def reject_doctor(id):
    guard = require_login("admin")
    if guard: return guard

    con = get_db()
    cur = con.cursor()
    cur.execute("DELETE FROM doctors WHERE id=%s AND status='pending'", (id,))
    con.commit()
    flash("Doctor registration rejected and removed.", "danger")
    return redirect("/pending_doctors")

# ══════════════════════════════════════════════════════════════════════════════
# DOCTOR DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/doctor_dashboard")
def doctor_dashboard():
    guard = require_login("doctor")
    if guard: return guard

    con = get_db()
    cur = con.cursor()

    doctor_id = session.get("doctor_id")

    # Doctor's own profile
    cur.execute("SELECT * FROM doctors WHERE id=%s", (doctor_id,))
    doctor_info = cur.fetchone()

    # Records added by this doctor
    cur.execute(
        "SELECT COUNT(*) FROM records WHERE doctor_name=%s",
        (session.get("user"),)
    )
    my_records_count = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM workers")
    total_workers = cur.fetchone()[0]

    # Recent records by this doctor
    cur.execute("""
        SELECT r.*, w.name
        FROM records r
        LEFT JOIN workers w ON r.worker_id = w.id
        WHERE r.doctor_name=%s
        ORDER BY r.visit_date DESC LIMIT 8
    """, (session.get("user"),))
    recent_records = cur.fetchall()

    return render_template(
        "doctor_dashboard.html",
        doctor_info=doctor_info,
        my_records_count=my_records_count,
        total_workers=total_workers,
        recent_records=recent_records
    )

# ══════════════════════════════════════════════════════════════════════════════
# WORKER DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/worker_dashboard")
def worker_dashboard():
    guard = require_login("worker")
    if guard: return guard

    con = get_db()
    cur = con.cursor()

    worker_id   = session.get("worker_id")
    worker_info = None
    total_records = 0
    total_reports = 0

    if worker_id:
        cur.execute("SELECT * FROM workers WHERE id=%s", (worker_id,))
        worker_info = cur.fetchone()

        cur.execute("SELECT COUNT(*) FROM records WHERE worker_id=%s", (worker_id,))
        total_records = cur.fetchone()[0]

        cur.execute("SELECT COUNT(*) FROM reports WHERE worker_id=%s", (worker_id,))
        total_reports = cur.fetchone()[0]

    return render_template(
        "worker_dashboard.html",
        worker_info=worker_info,
        total_records=total_records,
        total_reports=total_reports
    )

# ══════════════════════════════════════════════════════════════════════════════
# MY RECORDS  (worker)
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/my_records")
def my_records():
    guard = require_login("worker")
    if guard: return guard

    worker_id = session.get("worker_id")
    records   = []

    if worker_id:
        con = get_db()
        cur = con.cursor()
        cur.execute(
            "SELECT * FROM records WHERE worker_id=%s ORDER BY visit_date DESC",
            (worker_id,)
        )
        records = cur.fetchall()

    return render_template("my_records.html", records=records)

# ══════════════════════════════════════════════════════════════════════════════
# MY REPORTS  (worker)
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/my_reports")
def my_reports():
    guard = require_login("worker")
    if guard: return guard

    worker_id = session.get("worker_id")
    reports   = []

    if worker_id:
        con = get_db()
        cur = con.cursor()
        cur.execute(
            "SELECT id, filename FROM reports WHERE worker_id=%s",
            (worker_id,)
        )
        reports = cur.fetchall()

    return render_template("my_reports.html", reports=reports)

# ══════════════════════════════════════════════════════════════════════════════
# ALL WORKERS  (admin)
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/all_workers")
def all_workers():
    guard = require_login("admin")
    if guard: return guard

    con = get_db()
    cur = con.cursor()
    cur.execute("SELECT * FROM workers ORDER BY id DESC")
    workers = cur.fetchall()

    return render_template("all_workers.html", workers=workers)

# ══════════════════════════════════════════════════════════════════════════════
# ADD WORKER  (admin)
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/add_worker", methods=["GET", "POST"])
def add_worker():
    guard = require_login("admin")
    if guard: return guard

    if request.method == "POST":
        name            = request.form["name"].strip()
        gender          = request.form["gender"]
        age             = request.form["age"]
        phone           = request.form["phone"].strip()
        region          = request.form["region"]
        origin_state    = request.form.get("origin_state", "")
        worker_password = request.form["worker_password"].strip()

        if len(phone) != 10 or not phone.isdigit():
            flash("Phone number must be exactly 10 digits.", "danger")
            return redirect("/add_worker")

        if len(worker_password) < 4:
            flash("Password must be at least 4 characters.", "danger")
            return redirect("/add_worker")

        con = get_db()
        cur = con.cursor()

        cur.execute("SELECT id FROM workers WHERE phone=%s", (phone,))
        if cur.fetchone():
            flash("A worker with this phone number already exists.", "danger")
            return redirect("/add_worker")

        cur.execute(
            """INSERT INTO workers(name, gender, age, phone, region, origin_state, worker_password)
               VALUES(%s,%s,%s,%s,%s,%s,%s)""",
            (name, gender, age, phone, region, origin_state, worker_password)
        )
        con.commit()
        flash(f"Worker '{name}' registered! Login: phone {phone}, password as set.", "success")
        return redirect("/add_worker")

    return render_template("add_worker.html")

# ══════════════════════════════════════════════════════════════════════════════
# ADD MEDICAL RECORD  (doctor)
# Doctor name auto-filled from session — no manual input needed
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/add_record", methods=["GET", "POST"])
def add_record():
    guard = require_login("doctor")
    if guard: return guard

    con = get_db()
    cur = con.cursor()
    cur.execute("SELECT id, name FROM workers ORDER BY name")
    workers = cur.fetchall()

    if request.method == "POST":
        worker_id  = request.form["worker_id"]
        symptoms   = request.form["symptoms"]
        diagnosis  = request.form["diagnosis"]
        medicines  = request.form["medicines"]
        # Doctor name comes from session — not from form input
        doctor_name = session.get("user")

        cur.execute(
            """INSERT INTO records(worker_id, symptoms, diagnosis, medicines, doctor_name, visit_date)
               VALUES(%s,%s,%s,%s,%s,CURDATE())""",
            (worker_id, symptoms, diagnosis, medicines, doctor_name)
        )
        con.commit()
        flash("Medical record added successfully!", "success")
        return redirect("/add_record")

    return render_template("add_record.html", workers=workers, doctor_name=session.get("user"))

# ══════════════════════════════════════════════════════════════════════════════
# SEARCH  (admin + doctor)
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/search", methods=["GET", "POST"])
def search():
    guard = require_login()
    if guard: return guard

    if session.get("role") == "worker":
        return redirect("/worker_dashboard")

    data    = []
    keyword = None

    if request.method == "POST":
        keyword = request.form["keyword"].strip()
        con = get_db()
        cur = con.cursor()
        cur.execute(
            """SELECT * FROM workers
               WHERE CAST(id AS CHAR) LIKE %s
               OR name LIKE %s
               OR phone LIKE %s""",
            ('%'+keyword+'%', '%'+keyword+'%', '%'+keyword+'%')
        )
        data = cur.fetchall()

    return render_template("search.html", data=data, keyword=keyword)

# ══════════════════════════════════════════════════════════════════════════════
# VIEW RECORDS
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/records/<int:id>")
def view_records(id):
    guard = require_login()
    if guard: return guard

    if session.get("role") == "worker" and session.get("worker_id") != id:
        flash("Access denied.", "danger")
        return redirect("/worker_dashboard")

    con = get_db()
    cur = con.cursor()

    cur.execute("SELECT * FROM workers WHERE id=%s", (id,))
    worker = cur.fetchone()

    cur.execute(
        "SELECT * FROM records WHERE worker_id=%s ORDER BY visit_date DESC",
        (id,)
    )
    records = cur.fetchall()

    return render_template("view_records.html", records=records, worker=worker)

# ══════════════════════════════════════════════════════════════════════════════
# FILE UPLOAD  (admin)
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/upload", methods=["GET", "POST"])
def upload():
    guard = require_login("admin")
    if guard: return guard

    con = get_db()
    cur = con.cursor()

    cur.execute("SELECT id, name FROM workers ORDER BY name")
    workers = cur.fetchall()

    cur.execute("""
        SELECT r.id, r.worker_id, w.name, r.filename
        FROM reports r
        LEFT JOIN workers w ON r.worker_id = w.id
        ORDER BY r.id DESC LIMIT 10
    """)
    recent_uploads = cur.fetchall()

    if request.method == "POST":
        worker_id = request.form["worker_id"]
        file      = request.files["file"]

        if file.filename == "":
            flash("Please select a file to upload.", "danger")
            return redirect("/upload")

        allowed = {"pdf", "jpg", "jpeg", "png", "docx"}
        ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
        if ext not in allowed:
            flash("File type not allowed. Upload PDF, JPG, PNG, or DOCX.", "danger")
            return redirect("/upload")

        filepath = os.path.join(app.config["UPLOAD_FOLDER"], file.filename)
        file.save(filepath)

        cur.execute(
            "INSERT INTO reports(worker_id, filename) VALUES(%s,%s)",
            (worker_id, file.filename)
        )
        con.commit()
        flash("Report uploaded successfully!", "success")
        return redirect("/upload")

    return render_template("upload.html", workers=workers, recent_uploads=recent_uploads)

# ══════════════════════════════════════════════════════════════════════════════
# DOWNLOAD
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/download/<filename>")
def download(filename):
    guard = require_login()
    if guard: return guard
    return send_from_directory(app.config["UPLOAD_FOLDER"], filename, as_attachment=True)

# ══════════════════════════════════════════════════════════════════════════════
# ANALYTICS  (admin)
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/analytics")
def analytics():
    guard = require_login("admin")
    if guard: return guard

    con = get_db()
    cur = con.cursor()

    cur.execute("SELECT COUNT(*) FROM workers")
    total_workers = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM records")
    total_records = cur.fetchone()[0]

    cur.execute("SELECT COUNT(DISTINCT diagnosis) FROM records")
    unique_diagnoses = cur.fetchone()[0]

    cur.execute("""
        SELECT COUNT(*) FROM records
        WHERE LOWER(diagnosis) LIKE '%fever%'
           OR LOWER(diagnosis) LIKE '%dengue%'
           OR LOWER(diagnosis) LIKE '%malaria%'
           OR LOWER(diagnosis) LIKE '%typhoid%'
           OR LOWER(diagnosis) LIKE '%covid%'
           OR LOWER(diagnosis) LIKE '%tuberculosis%'
           OR LOWER(diagnosis) LIKE '%tb%'
    """)
    infectious_cases = cur.fetchone()[0]

    cur.execute("SELECT region, COUNT(*) AS cnt FROM workers GROUP BY region ORDER BY cnt DESC")
    region_data = cur.fetchall()
    max_region  = max((r[1] for r in region_data), default=1)

    cur.execute("""
        SELECT diagnosis, COUNT(*) AS cnt
        FROM records GROUP BY diagnosis ORDER BY cnt DESC LIMIT 8
    """)
    diagnosis_data = cur.fetchall()
    max_diagnosis  = max((d[1] for d in diagnosis_data), default=1)

    cur.execute("SELECT gender, COUNT(*) FROM workers GROUP BY gender")
    gender_data = cur.fetchall()

    cur.execute("SELECT age FROM workers")
    ages = [row[0] for row in cur.fetchall()]
    age_buckets = {"18-25": 0, "26-35": 0, "36-45": 0, "46-55": 0, "55+": 0}
    for a in ages:
        if   a <= 25: age_buckets["18-25"] += 1
        elif a <= 35: age_buckets["26-35"] += 1
        elif a <= 45: age_buckets["36-45"] += 1
        elif a <= 55: age_buckets["46-55"] += 1
        else:         age_buckets["55+"]   += 1
    age_data = list(age_buckets.items())
    max_age  = max((v for _, v in age_data), default=1)

    cur.execute("""
        SELECT w.name, w.region, r.diagnosis, r.doctor_name, r.visit_date
        FROM records r
        LEFT JOIN workers w ON r.worker_id = w.id
        ORDER BY r.visit_date DESC LIMIT 15
    """)
    recent_records = cur.fetchall()

    return render_template(
        "analytics.html",
        total_workers=total_workers,
        total_records=total_records,
        unique_diagnoses=unique_diagnoses,
        infectious_cases=infectious_cases,
        region_data=region_data,
        max_region=max_region,
        diagnosis_data=diagnosis_data,
        max_diagnosis=max_diagnosis,
        gender_data=gender_data,
        age_data=age_data,
        max_age=max_age,
        recent_records=recent_records
    )

# ══════════════════════════════════════════════════════════════════════════════
# DOWNLOAD DOCTORS EXCEL  (admin)
# Generates a fresh Excel from the live doctors table — always up to date.
# Newly approved self-registered doctors are included automatically.
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/download_doctors_excel")
def download_doctors_excel():
    guard = require_login("admin")
    if guard: return guard

    con = get_db()
    cur = con.cursor()

    # Fetch ALL approved doctors ordered by id
    cur.execute("""
        SELECT id, name, specialization, qualification, hospital,
               district, phone, email, experience, username, status, source
        FROM doctors
        WHERE status = 'approved'
        ORDER BY id
    """)
    doctors = cur.fetchall()

    # ── Build workbook ────────────────────────────────────────────────────────
    wb  = Workbook()
    ws  = wb.active
    ws.title = "Doctor Directory"

    HDR_BG = "1A56DB"
    HDR_FG = "FFFFFF"
    ALT_BG = "EFF6FF"

    thin   = Side(style="thin", color="D0D7E2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title row
    ws.merge_cells("A1:K1")
    t = ws["A1"]
    t.value     = "MigraHealth Kerala — Doctor Directory"
    t.font      = Font(name="Arial", bold=True, size=14, color=HDR_FG)
    t.fill      = PatternFill("solid", fgColor="0F172A")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    # Sub-title
    ws.merge_cells("A2:K2")
    s = ws["A2"]
    s.value     = f"Government of Kerala — Health Service Department | {len(doctors)} Approved Doctors | Generated on demand"
    s.font      = Font(name="Arial", size=10, color="64748B", italic=True)
    s.fill      = PatternFill("solid", fgColor="F1F5F9")
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # Header row
    headers = [
        "Doctor ID", "Full Name", "Specialization", "Qualification",
        "Hospital / Health Centre", "District", "Phone", "Email",
        "Experience (Yrs)", "Username", "Source"
    ]
    ws.row_dimensions[3].height = 26
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font      = Font(name="Arial", bold=True, size=11, color=HDR_FG)
        c.fill      = PatternFill("solid", fgColor=HDR_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = border

    # Data rows
    # doctors row: id(0) name(1) spec(2) qual(3) hosp(4) dist(5)
    #              phone(6) email(7) exp(8) username(9) status(10) source(11)
    for i, d in enumerate(doctors):
        row = i + 4
        source_label = "Pre-loaded" if d[11] == "dataset" else "Self-Registered"
        row_data = [
            f"D{d[0]:03d}", d[1], d[2], d[3], d[4],
            d[5], d[6], d[7], d[8], d[9], source_label
        ]
        bg = ALT_BG if i % 2 == 0 else "FFFFFF"
        ws.row_dimensions[row].height = 20

        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font      = Font(name="Arial", size=10,
                               bold=(col == 1),
                               color=("7C3AED" if col == 1 else "0F172A"))
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(
                horizontal="center" if col in (1, 7, 9) else "left",
                vertical="center",
                wrap_text=(col in (4, 5))
            )
            c.border = border

    # Column widths
    col_widths = [10, 26, 22, 18, 42, 20, 14, 38, 16, 22, 14]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:K{len(doctors) + 3}"

    # ── Stream to browser ─────────────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="Doctors_Dataset_MigraHealth_Kerala.xlsx"
    )


if __name__ == "__main__":
    app.run(debug=True)